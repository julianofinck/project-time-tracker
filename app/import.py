# pip install pandas
# pip install requests
# pip install openpyxl
# pip install Office365-REST-Python-Client
import pandas as pd
import requests
from office365.runtime.auth.authentication_context import AuthenticationContext
from office365.sharepoint.client_context import ClientContext
from openpyxl import load_workbook

# import psycopg2

# Authenticate with SharePoint (if necessary)

# Access the Excel file
excel_url = "hhttps://imgoffice.sharepoint.com/:x:/r/sites/CodexMeioAmbiente/_layouts/15/Doc.aspx?sourcedoc=%7B8E72189C-715D-4E9D-9679-2A2C9D0A2A8D%7D&file=G_ao_K_APONTAMENTOS_OPE_2023.xlsx&action=default&mobileredirect=true"
response = requests.get(excel_url)

# Load the Excel file
wb = load_workbook(filename=BytesIO(response.content))

# Extract data from the Excel file
data = []
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
