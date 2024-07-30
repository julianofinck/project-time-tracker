# pip install pandas
# pip install requests
# pip install openpyxl
# pip install Office365-REST-Python-Client
from io import BytesIO

import pandas as pd
import requests
from office365.runtime.auth.authentication_context import AuthenticationContext
from office365.sharepoint.client_context import ClientContext
from openpyxl import load_workbook

# import psycopg2

# Authenticate with SharePoint (if necessary)

# Access the Excel file
excel_url = "EXCEL_URL"
response = requests.get(excel_url)

# Load the Excel file
wb = load_workbook(filename=BytesIO(response.content))

# Extract data from the Excel file
data = []
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
